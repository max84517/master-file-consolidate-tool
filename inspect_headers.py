import sys
from pathlib import Path
from openpyxl import load_workbook

# Add src to path just in case, though might not be needed for this standalone script
sys.path.insert(0, r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\src")

source_dir = Path(r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\data\source\Peripheral")

if not source_dir.exists():
    print(f"Directory not found: {source_dir}")
    sys.exit(1)

for xf in sorted(source_dir.glob("*.xlsx")):
    try:
        # read_only=True is faster and safer for inspecting
        wb = load_workbook(xf, data_only=True, read_only=True)
    except Exception as e:
        print(f"{xf.name}: ERROR {e}")
        continue
    for sn in wb.sheetnames:
        ws = wb[sn]
        # Using iter_rows with read_only=True
        rows_iter = ws.iter_rows(values_only=True)
        try:
            # Skip first row and get the second row as header (as per script logic)
            next(rows_iter) 
            header_raw = next(rows_iter)
        except StopIteration:
            continue
            
        if not header_raw:
            continue
            
        headers = [str(h).replace("\n", " ").strip() if h is not None else "" for h in header_raw]
        suspect = [h for h in headers if h.lower() in ("odm cost", "rebate", "hp cost")]
        if suspect:
            print(f"\n{xf.name} | sheet={sn}")
            print(f"  Suspect cols: {suspect}")
            print(f"  All value cols: {[h for h in headers if any(k in h.lower() for k in ['hp cost','odm cost','rebate'])]}")
    wb.close()
