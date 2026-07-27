import sys
from pathlib import Path
from openpyxl import load_workbook

source_dir = Path(r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\data\source\Peripheral")

if not source_dir.exists():
    print(f"Directory not found: {source_dir}")
    sys.exit(1)

files = list(source_dir.glob("*.xlsx"))
print(f"Found {len(files)} excel files in {source_dir}")

for xf in sorted(files):
    print(f"Checking {xf.name}...")
    try:
        wb = load_workbook(xf, data_only=True, read_only=True)
    except Exception as e:
        print(f"  {xf.name}: ERROR {e}")
        continue
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            row1 = next(rows_iter)
            row2 = next(rows_iter)
        except StopIteration:
            continue
            
        # Try both row 1 and row 2 as headers to be sure
        for i, header_raw in enumerate([row1, row2]):
            if not header_raw:
                continue
            headers = [str(h).replace("\n", " ").strip() if h is not None else "" for h in header_raw]
            suspect = [h for h in headers if h.lower() in ("odm cost", "rebate", "hp cost")]
            if suspect:
                print(f"  MATCH: {xf.name} | sheet={sn} | row={i+1}")
                print(f"    Suspect cols: {suspect}")
                print(f"    All value cols: {[h for h in headers if any(k in h.lower() for k in ['hp cost','odm cost','rebate'])]}")
    wb.close()
