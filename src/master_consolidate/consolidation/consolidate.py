"""
Consolidation module.

Steps:
1. For each segment (NB, DT, Peripheral), collect all supplier Excels from
   data/source/<Segment>/.
2. For each Excel, open the sheet matching the chosen FY.
3. Clean the sheet:
   - Header is row 2.
   - Keep mandatory feature columns.
   - Drop rows where Platforms/Project is blank.
   - Keep value columns whose header contains any of the chosen keywords
     (HP Cost, ODM Cost, Rebate) — case-insensitive substring match.
   - Drop all other columns.
   - Replace \\n in column headers with space.
4. Stack all suppliers for a segment → save Consolidated Master price table_<Segment>_<Date>.xlsx
5. (Consolidate All) stack all three segment outputs → save Final Consolidated Master price table_<Date>.xlsx
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl import load_workbook, Workbook

# ── constants ────────────────────────────────────────────────────────────────

MANDATORY_COLS = [
    "Segment",
    "Category",
    "SPM (Project Owner)",
    "HP/ODM Part#",
    "Series",
    "Platforms/Project",
    "Product",
    "Size",
    "Product Type",
    "Color",
    "Other Description",
    "ODM (Regional Site)",
    "IncoTerm",
    "GTK Suppliers",
]

VALUE_KEYWORDS = {
    "HP Cost": "hp cost",
    "ODM Cost": "odm cost",
    "Rebate": "rebate",
}

_RE_NB_SHEET = re.compile(r"^fy(\d+)[cb]nb$")
_RE_OTHER_SHEET = re.compile(r"^fy(\d+)$")

# Value column names must follow "<keyword> <Month> <Year>" format,
# e.g. "HP Cost Nov 2025", "ODM Cost April 2026", "Rebate Jan 2024".
# Columns that match a keyword but lack a month+year suffix are rejected.
_RE_VALUE_COL = re.compile(
    r"^.+\s+(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec"
    r"|january|february|march|april|june|july|august|september|october|november|december)"
    r"\s+\d{4}$",
    re.IGNORECASE,
)

# ── HP Fiscal Year year-correction helpers ───────────────────────────────────
# HP FY starts Nov 1 and ends Oct 31 of the next calendar year.
# e.g. FY26 = Nov 2025 … Oct 2026
#   Nov, Dec → calendar year 2000+FY-1
#   Jan–Oct  → calendar year 2000+FY

_MONTH_NUM: dict[str, int] = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

_RE_EXTRACT_MONTH_YEAR = re.compile(
    r"^(.*?)\s+(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec"
    r"|january|february|march|april|june|july|august|september|october|november|december)"
    r"\s+(\d{4})$",
    re.IGNORECASE,
)


def _expected_year_for_month(month_str: str, fy_xx: str) -> int:
    """Return the expected calendar year for *month_str* in HP FY *fy_xx*."""
    fy_num = int(fy_xx)
    month_num = _MONTH_NUM.get(month_str.lower(), 0)
    return (2000 + fy_num - 1) if month_num >= 11 else (2000 + fy_num)


def _correct_value_col_year(header: str, fy_xx: str) -> tuple[str, bool]:
    """
    Check whether the year in a value-column header matches HP FY rules.
    Returns (corrected_header, was_corrected).
    If the header does not match the expected pattern, return it unchanged.
    """
    m = _RE_EXTRACT_MONTH_YEAR.match(header)
    if not m:
        return header, False
    prefix, month, year_str = m.group(1), m.group(2), m.group(3)
    expected = _expected_year_for_month(month, fy_xx)
    if int(year_str) == expected:
        return header, False
    return f"{prefix} {month} {expected}", True


# Known header typos across supplier files → canonical MANDATORY_COLS name.
# key: normalised (lowercase, no extra spaces), value: canonical display name.
_MANDATORY_ALIASES: dict[str, str] = {
    "incotrem": "IncoTerm",   # Darfon cNB sheets (e/r transposed)
}

# Lookup: canonical_lower → position in MANDATORY_COLS
_MAND_ORDER: dict[str, int] = {m.lower(): i for i, m in enumerate(MANDATORY_COLS)}


def _canonical_mandatory(raw: str) -> str | None:
    """
    Return the canonical MANDATORY_COLS name for *raw*, or None.
    Handles case differences and known typos via _MANDATORY_ALIASES.
    """
    normalized = raw.lower().strip()
    if normalized in _MANDATORY_ALIASES:
        return _MANDATORY_ALIASES[normalized]
    for mand in MANDATORY_COLS:
        if mand.lower() == normalized:
            return mand
    return None


def _norm(s: str) -> str:
    return s.lower().replace(" ", "")


# ── sheet matching ───────────────────────────────────────────────────────────

def _drop_unnamed_cols(ws) -> None:
    """Delete any column whose header cell (row 1) is blank or None."""
    to_delete = [
        col_idx
        for col_idx in range(ws.max_column, 0, -1)
        if not str(ws.cell(row=1, column=col_idx).value or "").strip()
    ]
    for col_idx in to_delete:
        ws.delete_cols(col_idx)


def _matching_sheets(wb: Workbook, segment: str, fy_xx: str) -> list[str]:
    """Return sheet names in wb that match the chosen FY for the given segment."""
    matches = []
    for sn in wb.sheetnames:
        n = _norm(sn)
        if segment == "NB":
            m = _RE_NB_SHEET.match(n)
            if m and m.group(1) == fy_xx:
                matches.append(sn)
        else:
            m = _RE_OTHER_SHEET.match(n)
            if m and m.group(1) == fy_xx:
                matches.append(sn)
    return matches


# ── sheet cleaning ───────────────────────────────────────────────────────────

def _clean_sheet(
    ws,
    value_keywords: list[str],
    fy_xx: str = "",
) -> tuple[list[str], list[list], list[tuple[str, str]]]:
    """
    Clean one worksheet and return (header, data_rows, corrections).
    header      — list of canonical column names (row 2 of source, normalised).
    data_rows   — list of rows (each a list aligned to header).
    corrections — list of (original_header, corrected_header) pairs where the
                  calendar year was adjusted to match HP FY rules.
    Returns ([], [], []) if no usable columns found.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 2:
        return [], [], []

    header_raw = list(all_rows[1])  # row 2 (0-indexed)

    # Normalise raw header cells: replace \n → space, strip
    raw_headers = [
        str(h).replace("\n", " ").strip() if h is not None else ""
        for h in header_raw
    ]

    kw_lower = [k.lower() for k in value_keywords]
    seen_mandatory: set[str] = set()

    keep_indices: list[int] = []
    canonical_names: list[str] = []
    is_value_col: list[bool] = []   # True = value col (None→0), False = feature col (None→"")
    corrections: list[tuple[str, str]] = []

    for i, h in enumerate(raw_headers):
        canon = _canonical_mandatory(h)
        if canon is not None:
            # Only keep first occurrence of each mandatory column
            if canon.lower() not in seen_mandatory:
                keep_indices.append(i)
                canonical_names.append(canon)
                is_value_col.append(False)
                seen_mandatory.add(canon.lower())
        elif any(kw in h.lower() for kw in kw_lower):
            if not _RE_VALUE_COL.match(h):
                continue  # e.g. bare "ODM Cost" or "Rebate" without month+year
            if fy_xx:
                corrected_h, was_corrected = _correct_value_col_year(h, fy_xx)
                if was_corrected:
                    corrections.append((h, corrected_h))
                h = corrected_h
            keep_indices.append(i)
            canonical_names.append(h)
            is_value_col.append(True)

    if not keep_indices:
        return [], [], []

    # Find Platforms/Project index (for blank-row filtering)
    try:
        pp_pos = [c.lower() for c in canonical_names].index("platforms/project")
    except ValueError:
        pp_pos = None

    data_rows: list[list] = []
    for row in all_rows[2:]:
        kept = []
        for pos, (i, is_val) in enumerate(zip(keep_indices, is_value_col)):
            raw = row[i] if i < len(row) else None
            if raw is None:
                kept.append(0 if is_val else "N/A")
            else:
                kept.append(raw)
        if pp_pos is not None:
            val = kept[pp_pos]
            if val == "" or val == "N/A" or (isinstance(val, str) and not val.strip()):
                continue
        data_rows.append(kept)

    return canonical_names, data_rows, corrections


# ── segment consolidation ────────────────────────────────────────────────────

def consolidate_segment(
    segment: str,
    source_dir: Path,
    fy: str,
    value_keywords: list[str],
    output_path: Path,
    log_fn=None,  # callable(str) | None — for real-time correction logging
) -> tuple[Path, list[dict]]:
    """
    Consolidate all supplier Excels for one segment into a single output Excel.
    Uses column-name alignment so sheets with different column sets or typos
    are handled correctly.

    Returns (output_file, corrections) where corrections is a list of dicts:
      {"segment", "supplier", "sheet", "original", "corrected"}
    """
    fy_xx = fy[2:]
    all_corrections: list[dict] = []

    # ── Pass 1: collect cleaned (header, rows) from every matching sheet ──
    all_data: list[tuple[list[str], list[list]]] = []
    for xf in sorted(source_dir.glob("*.xlsx")):
        supplier_name = xf.stem
        try:
            wb = load_workbook(xf, data_only=True)
        except Exception:
            continue
        for sn in _matching_sheets(wb, segment, fy_xx):
            header, rows, corrections = _clean_sheet(wb[sn], value_keywords, fy_xx)
            for orig, corr in corrections:
                info = {
                    "segment": segment,
                    "supplier": supplier_name,
                    "sheet": sn,
                    "original": orig,
                    "corrected": corr,
                }
                all_corrections.append(info)
                if log_fn:
                    log_fn(
                        f"  [CORRECTED] [{segment}] {supplier_name} ({sn}): "
                        f'"{orig}" → "{corr}"'
                    )
            if header and rows:
                all_data.append((header, rows))
        wb.close()

    # ── Pass 2: build unified header ──
    # Mandatory cols: collect unique canonical names, sort by MANDATORY_COLS order
    mand_seen: dict[str, str] = {}   # lower → display
    value_seen: dict[str, str] = {}  # lower → display  (preserves first-seen casing)
    for header, _ in all_data:
        for col in header:
            if _canonical_mandatory(col) is not None:
                if col.lower() not in mand_seen:
                    mand_seen[col.lower()] = col
            else:
                if col.lower() not in value_seen:
                    value_seen[col.lower()] = col

    sorted_mand = sorted(
        mand_seen.values(),
        key=lambda c: _MAND_ORDER.get(c.lower(), 999),
    )
    unified_header: list[str] = sorted_mand + list(value_seen.values())
    unified_lower = [c.lower() for c in unified_header]

    # ── Pass 3: write aligned rows ──
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = fy
    ws_out.append(unified_header)

    for header, rows in all_data:
        # Map each column in this sheet → index in unified_header
        sheet_to_unified: dict[int, int] = {}
        for sheet_pos, col in enumerate(header):
            try:
                unified_pos = unified_lower.index(col.lower())
                sheet_to_unified[sheet_pos] = unified_pos
            except ValueError:
                pass  # col not in unified (shouldn't happen)

    # Pre-compute which unified columns are value cols (for default fill)
    unified_is_value = [
        _canonical_mandatory(c) is None
        for c in unified_header
    ]

    # ── Pass 3: build aligned rows (deduplicated) ──
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = fy
    ws_out.append(unified_header)

    seen_rows: set[tuple] = set()
    for header, rows in all_data:
        # Map each column in this sheet → index in unified_header
        sheet_to_unified: dict[int, int] = {}
        for sheet_pos, col in enumerate(header):
            try:
                unified_pos = unified_lower.index(col.lower())
                sheet_to_unified[sheet_pos] = unified_pos
            except ValueError:
                pass  # col not in unified (shouldn't happen)

        for row in rows:
            # Default: feature cols → "N/A", value cols → 0
            aligned = [0 if is_val else "N/A" for is_val in unified_is_value]
            for sheet_pos, unified_pos in sheet_to_unified.items():
                if sheet_pos < len(row):
                    aligned[unified_pos] = row[sheet_pos]
            key = tuple(aligned)
            if key in seen_rows:
                continue
            seen_rows.add(key)
            ws_out.append(aligned)

    today = date.today().strftime("%Y%m%d")
    filename = f"Consolidated Master price table_{segment}_{today}.xlsx"
    out_file = output_path / filename
    output_path.mkdir(parents=True, exist_ok=True)
    _drop_unnamed_cols(ws_out)
    wb_out.save(out_file)
    return out_file, all_corrections


# ── all segments consolidation ───────────────────────────────────────────────

def consolidate_all(
    segment_files: dict[str, Path],  # {"NB": Path, "DT": Path, "Peripheral": Path}
    output_path: Path,
) -> Path:
    """
    Merge the three segment-level consolidated Excels into one final file.
    """
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "All"

    header_written = False
    all_data_rows: list[list] = []

    for seg in ("NB", "DT", "Peripheral"):
        seg_file = segment_files.get(seg)
        if seg_file is None or not seg_file.exists():
            continue
        try:
            wb = load_workbook(seg_file, data_only=True)
        except Exception:
            continue
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            wb.close()
            continue
        if not header_written:
            ws_out.append(list(rows[0]))
            header_written = True
            data_rows = rows[1:]
        else:
            data_rows = rows[1:]
        for row in data_rows:
            all_data_rows.append(list(row))
        wb.close()

    seen_rows: set[tuple] = set()
    for row in all_data_rows:
        key = tuple(row)
        if key in seen_rows:
            continue
        seen_rows.add(key)
        ws_out.append(row)

    today = date.today().strftime("%Y%m%d")
    filename = f"Final Consolidated Master price table_{today}.xlsx"
    out_file = output_path / filename
    output_path.mkdir(parents=True, exist_ok=True)
    _drop_unnamed_cols(ws_out)
    wb_out.save(out_file)
    return out_file
