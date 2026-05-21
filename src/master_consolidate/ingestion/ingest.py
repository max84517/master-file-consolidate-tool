"""
Ingestion module — scans source folders for supplier Excels, filters sheets,
and copies them to data/source/<Segment>/.

Sheet-matching rules:
  NB  : sheets whose name (normalised: lowercase, no spaces) matches
        /^fy[0-9]+[cb]nb$/  e.g. "FY26 cNB", "fy26bNB", "Fy 26 C NB"
  DT/Peripheral: sheets whose name (normalised) matches /^fy[0-9]+$/
                  e.g. "FY26", "fy 26"

The newest Excel in each supplier folder is selected by file mtime.
"""
import re
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl import load_workbook


# ── normalisation helpers ────────────────────────────────────────────────────

_RE_NB_SHEET = re.compile(r"^fy(\d+)[cb]nb$")
_RE_OTHER_SHEET = re.compile(r"^fy(\d+)$")


def _norm(name: str) -> str:
    """Lowercase, strip all spaces."""
    return name.lower().replace(" ", "")


def _canonical_nb(raw: str) -> str:
    """Return canonical sheet name FYXX cNB or FYXX bNB."""
    n = _norm(raw)
    m = _RE_NB_SHEET.match(n)
    if not m:
        raise ValueError(f"Not a valid NB sheet: {raw!r}")
    xx = m.group(1)
    letter = "c" if "cnb" in n else "b"
    return f"FY{xx} {letter}NB"


def _canonical_other(raw: str) -> str:
    """Return canonical sheet name FYXX."""
    n = _norm(raw)
    m = _RE_OTHER_SHEET.match(n)
    if not m:
        raise ValueError(f"Not a valid FY sheet: {raw!r}")
    return f"FY{m.group(1)}"


def is_nb_sheet(name: str) -> bool:
    return bool(_RE_NB_SHEET.match(_norm(name)))


def is_other_sheet(name: str) -> bool:
    return bool(_RE_OTHER_SHEET.match(_norm(name)))


# ── file discovery ───────────────────────────────────────────────────────────

def _latest_excel(folder: Path) -> Path | None:
    """Return the most-recently modified .xlsx/.xls file in *folder*."""
    excels = [f for f in folder.iterdir()
              if f.is_file() and f.suffix.lower() in (".xlsx", ".xls", ".xlsm")]
    if not excels:
        return None
    return max(excels, key=lambda f: f.stat().st_mtime)


def discover_suppliers(segment_root: Path) -> dict[str, Path]:
    """
    Returns {supplier_name: latest_excel_path} for each supplier sub-folder.
    segment_root is e.g. <NB KB folder>/Master price table_NB
    """
    result: dict[str, Path] = {}
    if not segment_root.exists():
        return result
    for item in sorted(segment_root.iterdir()):
        if not item.is_dir():
            continue
        # Skip "Consolidated Master price table_*" output folders
        if item.name.lower().startswith("consolidated master price table"):
            continue
        latest = _latest_excel(item)
        if latest:
            result[item.name] = latest
    return result


# ── ingestion ────────────────────────────────────────────────────────────────

def ingest_segment(
    segment: str,          # "NB" | "DT" | "Peripheral"
    source_root: Path,     # top-level folder chosen by user (e.g. NB KB path)
    dest_dir: Path,        # data/source/<Segment>
) -> dict[str, list[str]]:
    """
    Reads each supplier's latest Excel, copies only the relevant sheets
    to dest_dir/<supplier_filename>.xlsx.

    Returns {supplier_folder_name: [canonical_sheet_names_found]}.
    """
    # Locate the Master price table sub-folder
    # It might be named "Master price table_NB" or similar; we search case-insensitively.
    master_folder = _find_master_folder(source_root, segment)
    if master_folder is None:
        raise FileNotFoundError(
            f"Cannot find 'Master price table_{segment}' folder inside: {source_root}"
        )

    suppliers = discover_suppliers(master_folder)
    summary: dict[str, list[str]] = {}

    dest_dir.mkdir(parents=True, exist_ok=True)

    for supplier_folder_name, excel_path in suppliers.items():
        try:
            copied_sheets = _copy_sheets(segment, excel_path, dest_dir)
            summary[supplier_folder_name] = copied_sheets
        except Exception as exc:
            summary[supplier_folder_name] = [f"ERROR: {exc}"]

    return summary


def _find_master_folder(root: Path, segment: str) -> Path | None:
    """Find the Master price table folder for the given segment (case-insensitive)."""
    target = f"master price table_{segment}".lower()
    for item in root.iterdir():
        if item.is_dir() and item.name.lower() == target:
            return item
    # fallback: any dir starting with "master price table"
    for item in root.iterdir():
        if item.is_dir() and item.name.lower().startswith("master price table"):
            return item
    return None


def _copy_sheets(segment: str, src_path: Path, dest_dir: Path) -> list[str]:
    """
    Open src_path, keep only the relevant sheets, save to dest_dir.
    Returns list of canonical sheet names saved.
    """
    wb_src = load_workbook(src_path, data_only=True)
    wb_dest = openpyxl.Workbook()
    # Remove default sheet
    for s in wb_dest.sheetnames:
        del wb_dest[s]

    canonical_names: list[str] = []

    for sheet_name in wb_src.sheetnames:
        if segment == "NB":
            if not is_nb_sheet(sheet_name):
                continue
            canonical = _canonical_nb(sheet_name)
        else:
            if not is_other_sheet(sheet_name):
                continue
            canonical = _canonical_other(sheet_name)

        ws_src = wb_src[sheet_name]
        ws_dest = wb_dest.create_sheet(title=canonical)

        # Build a map: (row, col) → anchor value for every secondary cell in a
        # merged range.  This prevents column offsets when a header cell is
        # horizontally merged (e.g. "IncoTerm" spanning two columns).
        merge_fill: dict[tuple[int, int], object] = {}
        for rng in ws_src.merged_cells.ranges:
            anchor_val = ws_src.cell(
                row=rng.min_row, column=rng.min_col
            ).value
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if r != rng.min_row or c != rng.min_col:
                        merge_fill[(r, c)] = anchor_val

        for row in ws_src.iter_rows():
            for cell in row:
                value = merge_fill.get((cell.row, cell.column), cell.value)
                ws_dest.cell(row=cell.row, column=cell.column, value=value)

        canonical_names.append(canonical)

    if not canonical_names:
        return []

    out_path = dest_dir / src_path.name
    wb_dest.save(out_path)
    return canonical_names


# ── available FY years across all ingested files ─────────────────────────────

def available_fy_years(source_base: Path) -> list[str]:
    """
    Scan data/source/{NB,DT,Peripheral} and collect all FY years found in
    sheet names. Returns sorted list e.g. ["FY25", "FY26"].
    """
    fy_set: set[str] = set()
    for seg in ("NB", "DT", "Peripheral"):
        seg_dir = source_base / seg
        if not seg_dir.exists():
            continue
        for xf in seg_dir.glob("*.xlsx"):
            try:
                wb = load_workbook(xf, read_only=True, data_only=True)
                for sn in wb.sheetnames:
                    n = _norm(sn)
                    m_nb = _RE_NB_SHEET.match(n)
                    m_ot = _RE_OTHER_SHEET.match(n)
                    if m_nb:
                        fy_set.add(f"FY{m_nb.group(1)}")
                    elif m_ot:
                        fy_set.add(f"FY{m_ot.group(1)}")
                wb.close()
            except Exception:
                pass
    return sorted(fy_set, key=lambda s: int(s[2:]))


def check_fy_coverage(
    source_base: Path,
    fy: str,           # e.g. "FY26"
) -> dict[str, dict[str, bool]]:
    """
    Returns {segment: {supplier_filename: has_sheet}}.
    """
    xx = fy[2:]  # e.g. "26"
    result: dict[str, dict[str, bool]] = {}
    for seg in ("NB", "DT", "Peripheral"):
        seg_dir = source_base / seg
        seg_result: dict[str, bool] = {}
        if not seg_dir.exists():
            result[seg] = seg_result
            continue
        for xf in seg_dir.glob("*.xlsx"):
            try:
                wb = load_workbook(xf, read_only=True, data_only=True)
                found = False
                for sn in wb.sheetnames:
                    n = _norm(sn)
                    if seg == "NB":
                        if _RE_NB_SHEET.match(n) and _RE_NB_SHEET.match(n).group(1) == xx:
                            found = True
                            break
                    else:
                        if _RE_OTHER_SHEET.match(n) and _RE_OTHER_SHEET.match(n).group(1) == xx:
                            found = True
                            break
                wb.close()
                seg_result[xf.name] = found
            except Exception:
                seg_result[xf.name] = False
        result[seg] = seg_result
    return result
