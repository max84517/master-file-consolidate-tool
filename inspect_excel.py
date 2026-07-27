from pathlib import Path
from openpyxl import load_workbook

xf = Path(r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\data\source\Peripheral\Consolidated Master price table_Peripheral_20260320.xlsx")

if not xf.exists():
    print(f"File not found: {xf}")
else:
    wb = load_workbook(xf, data_only=True)
    print("Sheet names:", wb.sheetnames)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        print(f"\nSheet: {sn}, total rows: {len(rows)}")
        if rows:
            print(f"  Row 1: {rows[0][:20]}")
        if len(rows) >= 2:
            print(f"  Row 2: {rows[1][:20]}")
    wb.close()
