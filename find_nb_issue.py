from pathlib import Path
from openpyxl import load_workbook

source_dir = Path(r"data/source/NB")

for xf in sorted(source_dir.glob("*.xlsx")):
    try:
        wb = load_workbook(xf, data_only=True)
    except Exception as e:
        print(f"{xf.name}: ERROR {e}")
        continue
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        # Check both row 1 (index 0) and row 2 (index 1) as header
        for row_idx in [0, 1]:
            header_raw = rows[row_idx]
            headers = [str(h).replace("\n", " ").strip() if h is not None else "" for h in header_raw]
            value_cols = [h for h in headers if any(k in h.lower() for k in ["hp cost", "odm cost", "rebate"])]
            standalone = [h for h in value_cols if h.lower() in ("odm cost", "rebate", "hp cost")]
            if standalone:
                print(f"\n*** FOUND *** {xf.name} | sheet={sn} | header_row={row_idx+1}")
                print(f"  Standalone cols: {standalone}")
                print(f"  All value cols: {value_cols}")
    wb.close()
