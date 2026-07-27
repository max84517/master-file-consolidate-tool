import sys
from pathlib import Path
from openpyxl import load_workbook

source_dir = Path(r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\data\source\Peripheral")

for xf in sorted(source_dir.glob("*.xlsx")):
    print(f"\n--- {xf.name} ---")
    try:
        wb = load_workbook(xf, data_only=True, read_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                next(rows_iter) # Row 1
                row2 = next(rows_iter) # Row 2
                if row2:
                    headers = [str(h).replace("\n", " ").strip() if h is not None else "" for h in row2]
                    # Filter for anything that looks like cost or rebate
                    interesting = [h for h in headers if any(k in h.lower() for k in ["cost", "rebate"])]
                    if interesting:
                        print(f"Sheet: {sn}")
                        print(f"Headers containing 'cost' or 'rebate': {interesting}")
            except StopIteration:
                continue
        wb.close()
    except Exception as e:
        print(f"Error {xf.name}: {e}")
