from pathlib import Path
from openpyxl import load_workbook

files = [
    Path(r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\data\result_by_segment\Peripheral\Consolidated Master price table_Peripheral_20260521.xlsx"),
    Path(r"c:\Users\LiMa595\OneDrive - HP Inc\Input Device\17. Automation Projects\master-file-consolidate-tool\data\consolidate_all\Final Consolidated Master price table_20260521.xlsx"),
]

for xf in files:
    if not xf.exists():
        print(f"File not found: {xf}")
        continue
    wb = load_workbook(xf, data_only=True)
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        print(f"\n=== {xf.name} | sheet={sn} ===")
        for i, h in enumerate(header):
            print(f"  col {i:3d}: {repr(h)}")
    wb.close()
